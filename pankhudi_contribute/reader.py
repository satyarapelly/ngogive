from __future__ import annotations

import csv
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable

from .models import WorkbookProject, parse_date

REQUIRED_COLUMNS = {
    "Project UID": "project_uid",
    "Project ID": "project_id",
    "Complexity": "complexity",
    "Expected End Date": "expected_end_date",
    "Likely Contribute Button State (Derived)": "readiness_state",
    "Recommended Action": "recommended_action",
}

class WorkbookReadError(RuntimeError):
    pass

def read_projects(path: str | Path, sheet: str = "Ready to Contribute") -> list[WorkbookProject]:
    source = Path(path)
    if not source.exists():
        raise WorkbookReadError(f"Input file not found: {source}")
    if source.suffix.lower() == ".csv":
        return _rows_to_projects(_read_csv(source))
    if source.suffix.lower() in {".xlsx", ".xlsm"}:
        return _rows_to_projects(_read_xlsx(source, sheet))
    raise WorkbookReadError(f"Unsupported input file type: {source.suffix}")

def _read_csv(path: Path) -> list[dict[str, Any]]:
    with path.open(newline="", encoding="utf-8-sig") as fh:
        return list(csv.DictReader(fh))

def _read_xlsx(path: Path, sheet: str) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as exc:
        raise WorkbookReadError("openpyxl is required to read .xlsx files. Run: python -m pip install -e .") from exc
    wb = load_workbook(path, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        raise WorkbookReadError(f"Sheet not found: {sheet}. Available sheets: {', '.join(wb.sheetnames)}")
    ws = wb[sheet]
    rows = ws.iter_rows(values_only=True)
    try:
        headers = [str(v).strip() if v is not None else "" for v in next(rows)]
    except StopIteration as exc:
        raise WorkbookReadError(f"Sheet is empty: {sheet}") from exc
    return [dict(zip(headers, values)) for values in rows if any(v not in (None, "") for v in values)]

def _rows_to_projects(rows: Iterable[dict[str, Any]]) -> list[WorkbookProject]:
    projects: list[WorkbookProject] = []
    missing: set[str] | None = None
    for row in rows:
        if missing is None:
            missing = set(REQUIRED_COLUMNS) - set(row)
            if missing:
                raise WorkbookReadError(f"Missing required columns: {', '.join(sorted(missing))}")
        uid = str(row.get("Project UID") or "").strip()
        if not uid:
            continue
        projects.append(WorkbookProject(
            project_uid=uid,
            project_id=_int_or_none(row.get("Project ID")),
            complexity=str(row.get("Complexity") or "").strip(),
            expected_end_date=_coerce_date(row.get("Expected End Date")),
            readiness_state=str(row.get("Likely Contribute Button State (Derived)") or "").strip(),
            recommended_action=str(row.get("Recommended Action") or "").strip(),
        ))
    return projects

def _int_or_none(value: Any) -> int | None:
    if value in (None, ""):
        return None
    return int(float(str(value).strip()))

def _coerce_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    return parse_date(value)
